"""
schemas/needlist.py
Schema for the needlist data of the project."""

from pathlib import Path
from typing import List, Optional, Self

import openpyxl
from pydantic import BaseModel, Field, field_validator, model_validator


class PathInput(BaseModel):

    folder_path: str
    excel_path: str

    @field_validator("folder_path")
    @classmethod
    def check_folder_path(cls, v: str):
        if not Path(v).exists():
            raise ValueError("Folder does not exist.")
        return v

    @field_validator("excel_path")
    @classmethod
    def check_excel_path(cls, v: str):
        if not Path(v).exists() or not v.endswith((".xls", ".xlsx")):
            raise ValueError(
                f"Excel file does not exist or is not a valid Excel file: {v}"
            )
        return v


class ExcelInput(BaseModel):
    sheet_name: str
    header_row: int = Field(default=1, ge=1)


class ColumnInput(BaseModel):
    column_names: List[str] = Field(default_factory=list)
    doc_no_column_name: str = ""


class NeedlistInput(BaseModel):

    path_input: PathInput
    excel_input: ExcelInput
    column_input: ColumnInput
    search_level: str = "file"

    @field_validator("search_level")
    @classmethod
    def check_search_level(cls, v: str):
        if v not in ["folder", "file"]:
            raise ValueError("Search level should be either 'folder' or 'file'.")

    @model_validator(mode="after")
    def check_sheet_names(self) -> Self:
        wb = None
        try:
            wb = openpyxl.load_workbook(self.path_input.excel_path, read_only=True)
            if isinstance(self.excel_input.sheet_name, int):
                if self.excel_input.sheet_name >= len(wb.sheetnames):
                    raise ValueError("Sheet index out of range.")
                ws = wb[wb.sheetnames[self.excel_input.sheet_name]]
            elif isinstance(self.excel_input.sheet_name, str):
                if self.excel_input.sheet_name not in wb.sheetnames:
                    raise ValueError("Sheet name does not exist in the excel file.")
                ws = wb[self.excel_input.sheet_name]
            else:
                raise ValueError("Sheet name must be a string or an integer.")

            row = list(
                ws.iter_rows(
                    min_row=self.excel_input.header_row,
                    max_row=self.excel_input.header_row,
                    values_only=True,
                )
            )[0]
            self.column_input.column_names = [
                str(cell) for cell in row if cell is not None
            ]

        except Exception as e:
            raise ValueError(f"Error reading Excel file: {e}")
        finally:
            if wb:
                wb.close()

        return self
